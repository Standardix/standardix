import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import unicodedata
from typing import Optional, List, Dict
import sys
import subprocess
import json
from urllib.parse import urljoin, urlparse
import requests
from bs4 import BeautifulSoup

from standardix_engine import standardix, read_table

# --------------------------------------------------
# Constantes pour le générateur de descriptions courtes
# --------------------------------------------------

SHEET_EN = "EN"
SHEET_FR = "FR"
RECIPE_SHEET_EN = "Recipe_EN"
RECIPE_SHEET_FR = "Recipe_FR"

COL_SKU = "sku"
COL_PRODUCT_TYPE = "product_type"
COL_SHORT_DESC = "short_description"
ORIG_DESC_COL = "Short Description"

EMPTY_MARKERS = {
    "",
    " ",
    None,
    "unmapped",
    "UNMAPPED",
    "undefined",
    "undefinied",
    "nan",
    "NaN",
    "NAN",
    "UNDEFINITE",   # valeur EN quand non trouvée dans le mapping
    "NON_MAPPÉ",    # valeur FR quand non trouvée dans le mapping
}

SOURCE_TYPE_MAP = {
    "attribute value": "ATTRIBUTE_VALUE",
    "valeur d'attribut": "ATTRIBUTE_VALUE",
    "attribute name": "ATTRIBUTE_NAME",
    "nom d'attribut": "ATTRIBUTE_NAME",
}

SEPARATOR_KEYWORDS = {
    "space": " ",
    "comma": ", ",
    "colon": ": ",
    "dash": " - ",
    "dot": ". ",
    "bullet": " • ",
    "espace": " ",
    "virgule": ", ",
    "virgule-espace": ", ",
    "deux_points": ": ",
    "tiret": " - ",
    "point": ". ",
    "puce": " • ",
    "'s": "'s ",
    "’s": "’s ",
}


def normalize_string(x: Optional[str]) -> str:
    if x is None:
        return ""
    return str(x).strip()


def strip_accents(s: str) -> str:
    s = normalize_string(s)
    nfkd = unicodedata.normalize("NFD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def canon_key(s: str) -> str:
    s = normalize_string(s).lower()
    return s.replace(" ", "").replace("_", "")


def canon_match_key(s: str) -> str:
    s = strip_accents(s).lower()
    return " ".join(s.split())


def is_empty_value(x: Optional[str]) -> bool:
    if x is None or pd.isna(x):
        return True
    s = str(x).strip()
    return s in EMPTY_MARKERS


def load_recipes(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """
    Valide et prépare le contenu d'une feuille de recettes
    déjà chargée en DataFrame.
    """
    required_cols = ["product_type", "order", "source_type", "attribute_name", "separator_after"]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Missing required column '{col}' in sheet '{sheet_name}'")

    # Colonne brand facultative
    if "brand" not in df.columns:
        df["brand"] = ""

    # Normalisation des colonnes texte
    for col in ["product_type", "brand", "source_type", "attribute_name"]:
        df[col] = df[col].apply(lambda v: normalize_string(v))

    return df


def resolve_source_type(raw: str) -> str:
    key = normalize_string(raw).lower()
    if key in SOURCE_TYPE_MAP:
        return SOURCE_TYPE_MAP[key]
    raise ValueError(f"Unknown source_type value: {raw!r}")


def resolve_separator(raw) -> str:
    if raw is None or pd.isna(raw):
        return ""
    raw_str = str(raw)
    key = raw_str.strip().lower()
    if key in ("", "nan", "none", "null"):
        return ""
    if key in SEPARATOR_KEYWORDS:
        return SEPARATOR_KEYWORDS[key]
    return raw_str


def build_attr_lookup(columns: List[str]) -> Dict[str, str]:
    """
    Construit une table de correspondance tolérante pour les noms de colonnes.
    """
    mapping: Dict[str, str] = {}
    suffixes = ["_standard_en", "_standard_fr"]

    for col in columns:
        if not col:
            continue
        col_clean = normalize_string(col)
        col_key = canon_key(col_clean)

        if col_key not in mapping:
            mapping[col_key] = col_clean

        col_lower = col_clean.lower()
        for suf in suffixes:
            if col_lower.endswith(suf):
                base = col_lower[: -len(suf)]
                base = base.rstrip("_")
                base_key = canon_key(base)
                if base_key and base_key not in mapping:
                    mapping[base_key] = col_clean

    return mapping


def get_attribute_value(row: pd.Series, attr_name: str, attr_lookup: Dict[str, str]) -> Optional[str]:
    if not attr_name:
        return None

    # Colonne présente telle quelle
    if attr_name in row.index:
        return row[attr_name]

    # Lookup canonique
    attr_key = canon_key(attr_name)
    if attr_key in attr_lookup:
        col = attr_lookup[attr_key]
        if col in row.index:
            return row[col]

    # Fallback : scan des colonnes avec clef canonique
    for col in row.index:
        if canon_key(col) == attr_key:
            return row[col]

    return None


def build_short_description_for_row(row: pd.Series, recipes: pd.DataFrame, attr_lookup: Dict[str, str]) -> str:
    """
    Construit la short description pour une ligne de produit.
    """
    product_type = normalize_string(row.get(COL_PRODUCT_TYPE, ""))
    if not product_type:
        return ""

    pt_key = canon_match_key(product_type)
    applicable = recipes[recipes["product_type"].apply(canon_match_key) == pt_key]
    if applicable.empty:
        return ""

    applicable = applicable.sort_values("order")
    parts = []

    for _, rec in applicable.iterrows():
        source_type_raw = rec["source_type"]
        attr_name = rec["attribute_name"]
        sep_raw = rec["separator_after"]

        try:
            source_type = resolve_source_type(source_type_raw)
        except ValueError:
            # Valeur de source_type inconnue → on ignore cette ligne de recette
            continue

        if source_type == "ATTRIBUTE_VALUE":
            value = get_attribute_value(row, attr_name, attr_lookup)
        elif source_type == "ATTRIBUTE_NAME":
            value = attr_name
        else:
            value = None

        if is_empty_value(value):
            continue

        value_str = str(value).strip()
        parts.append(value_str)

        sep = resolve_separator(sep_raw)
        if sep:
            parts.append(sep)

    if not parts:
        return ""

    # Évite de finir sur un séparateur seul
    last = str(parts[-1])
    if last.strip() in {",", ":", "-", ".", "•"}:
        parts = parts[:-1]

    text = "".join(str(p) for p in parts).strip()
    return text


def process_language(standardized_df: pd.DataFrame, recipes: pd.DataFrame, lang_label: str) -> pd.DataFrame:
    """
    Applique les recettes pour une langue donnée.
    """
    for col in [COL_SKU, COL_PRODUCT_TYPE]:
        if col not in standardized_df.columns:
            raise ValueError(f"Missing required column '{col}' in standardized data for language {lang_label}.")

    attr_lookup = build_attr_lookup(list(standardized_df.columns))

    out_rows = []
    for _, row in standardized_df.iterrows():
        short_desc = build_short_description_for_row(row, recipes, attr_lookup)
        row_out = {
            COL_SKU: row[COL_SKU],
            COL_PRODUCT_TYPE: row[COL_PRODUCT_TYPE],
        }
        if ORIG_DESC_COL in standardized_df.columns:
            row_out[ORIG_DESC_COL] = row[ORIG_DESC_COL]

        row_out[COL_SHORT_DESC] = short_desc
        out_rows.append(row_out)

    return pd.DataFrame(out_rows)


# --------------------------------------------------
# CONFIG STREAMLIT GÉNÉRALE
# --------------------------------------------------

# ------------------------------------------------------------
# Helpers (HTTP)
# ------------------------------------------------------------
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}


def http_get(url: str, timeout: int = 30) -> requests.Response:
    r = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
    r.raise_for_status()
    return r


def abs_url(base: str, href: str) -> str:
    return urljoin(base, href)


def normalize_image_url(u: str) -> str:
    return (u or "").replace("&amp;", "&").strip()


def image_dedupe_key(u: str) -> str:
    return (u or "").split("?", 1)[0].strip()


def best_from_srcset(srcset: str) -> str:
    if not srcset:
        return ""
    parts = [p.strip() for p in srcset.split(",") if p.strip()]
    if not parts:
        return ""
    return parts[-1].split(" ")[0].strip()


def img_tag_to_url(img, base_url: str) -> str:
    src = (img.get("src") or "").strip()
    if src:
        return abs_url(base_url, src)
    srcset = (img.get("srcset") or "").strip()
    if srcset:
        return abs_url(base_url, best_from_srcset(srcset))
    dsrc = (img.get("data-src") or "").strip()
    if dsrc:
        return abs_url(base_url, dsrc)
    dsrcset = (img.get("data-srcset") or "").strip()
    if dsrcset:
        return abs_url(base_url, best_from_srcset(dsrcset))
    return ""


# ------------------------------------------------------------
# Playwright helpers
# ------------------------------------------------------------
def ensure_playwright_chromium_installed():
    """
    Install chromium if Playwright is present but browsers are missing (Streamlit Cloud).
    Runs at most once per app session.
    """
    if st.session_state.get("_pw_chromium_installed_ok") is True:
        return
    if st.session_state.get("_pw_chromium_installed_attempted"):
        return

    st.session_state["_pw_chromium_installed_attempted"] = True
    subprocess.run(
        [sys.executable, "-m", "playwright", "install", "chromium"],
        check=True,
        capture_output=True,
        text=True,
    )
    st.session_state["_pw_chromium_installed_ok"] = True


def _launch_browser(p):
    """Launch Chromium, auto-install if missing."""
    try:
        return p.chromium.launch(headless=True)
    except Exception as e:
        msg = str(e)
        if "Executable doesn't exist" in msg or "playwright install" in msg:
            ensure_playwright_chromium_installed()
            return p.chromium.launch(headless=True)
        raise


def get_rendered_html_playwright(url: str, wait_ms: int = 600) -> str:
    """Render a page with Playwright and return its HTML."""
    from playwright.sync_api import sync_playwright  # type: ignore

    with sync_playwright() as p:
        browser = _launch_browser(p)
        context = browser.new_context(user_agent=HEADERS["User-Agent"])

        # Speed: block heavy resources (we only need DOM + attributes)
        def _route(route, request):
            if request.resource_type in ("image", "media", "font"):
                return route.abort()
            return route.continue_()

        context.route("**/*", _route)

        page = context.new_page()
        page.goto(url, wait_until="domcontentloaded", timeout=90000)
        page.wait_for_timeout(wait_ms)
        html = page.content()
        browser.close()
        return html


# ------------------------------------------------------------
# PLP extraction (product list)
# ------------------------------------------------------------
_PRODUCT_PATH_HINTS = (
    "/products/",  # Shopify + many
    "/product/",   # Rapha and others
)


def _same_origin(base_url: str, candidate_url: str) -> bool:
    try:
        b = urlparse(base_url)
        c = urlparse(candidate_url)
        return (b.scheme, b.netloc) == (c.scheme, c.netloc)
    except Exception:
        return False


def _looks_like_product_href(href: str) -> bool:
    h = (href or "").lower()
    if any(x in h for x in ("onetrust", "cookie", "consent", "privacy", "terms")):
        return False
    return any(p in h for p in _PRODUCT_PATH_HINTS)


def plp_products_requests(plp_url: str, max_products: int) -> list[dict]:
    """Extract product_name + product_url from a PLP HTML (no JS)."""
    base = f"{urlparse(plp_url).scheme}://{urlparse(plp_url).netloc}"
    html = http_get(plp_url).text
    soup = BeautifulSoup(html, "html.parser")

    products: list[dict] = []
    seen: set[str] = set()

    # Prefer links inside <main> if present (reduces cookie links)
    root = soup.find("main") or soup

    for a in root.find_all("a", href=True):
        href = (a.get("href") or "").strip()
        if not href:
            continue
        if not _looks_like_product_href(href):
            continue

        pdp = abs_url(base, href.split("?")[0])
        if not _same_origin(plp_url, pdp):
            continue
        if pdp in seen:
            continue

        # Ensure it's likely a product card: link has an image inside
        if not a.find("img"):
            continue

        name = a.get_text(" ", strip=True) or a.get("aria-label") or a.get("title") or ""
        name = (name or "").strip()

        seen.add(pdp)
        products.append({"product_name": name, "product_url": pdp})
        if len(products) >= max_products:
            break

    return products


def plp_products_playwright(plp_url: str, max_products: int) -> list[dict]:
    """Extract product_name + product_url from a PLP using rendered DOM (Playwright)."""
    base = f"{urlparse(plp_url).scheme}://{urlparse(plp_url).netloc}"
    from playwright.sync_api import sync_playwright  # type: ignore

    with sync_playwright() as p:
        browser = _launch_browser(p)
        context = browser.new_context(user_agent=HEADERS["User-Agent"])

        # Speed: block heavy resources
        def _route(route, request):
            if request.resource_type in ("image", "media", "font"):
                return route.abort()
            return route.continue_()

        context.route("**/*", _route)

        page = context.new_page()
        page.goto(plp_url, wait_until="domcontentloaded", timeout=90000)

        # Some PLPs load items on scroll
        for _ in range(5):
            page.mouse.wheel(0, 2200)
            page.wait_for_timeout(450)

        js = r"""
        () => {
          const bad = ['onetrust','cookie','consent','privacy','terms'];
          const hints = ['/products/','/product/'];

          function isBadHref(h){
            const s = (h||'').toLowerCase();
            return bad.some(b => s.includes(b));
          }
          function looksLikeProduct(h){
            const s = (h||'');
            return hints.some(x => s.includes(x));
          }

          const anchors = Array.from(document.querySelectorAll('main a[href], a[href]'));
          const out = [];
          for(const a of anchors){
            const href = a.getAttribute('href') || '';
            if(!href) continue;
            if(isBadHref(href)) continue;
            if(!looksLikeProduct(href)) continue;

            // must contain an image (product tile)
            if(!a.querySelector('img')) continue;

            const txt = (a.innerText || '').trim();
            const aria = a.getAttribute('aria-label') || '';
            const title = a.getAttribute('title') || '';
            const imgAlt = (a.querySelector('img')?.getAttribute('alt') || '').trim();

            out.push({
              href,
              text: txt,
              aria,
              title,
              imgAlt
            });
          }
          return out;
        }
        """
        items = page.evaluate(js) or []
        browser.close()

    products: list[dict] = []
    seen: set[str] = set()

    for it in items:
        href = (it.get("href") or "").strip()
        if not href:
            continue
        pdp = abs_url(base, href.split("?")[0])
        if not _same_origin(plp_url, pdp):
            continue
        if pdp in seen:
            continue

        # Prefer image alt for name (Rapha often has it), else aria/title/text
        name = (it.get("imgAlt") or "").strip() or (it.get("aria") or "").strip() or (it.get("title") or "").strip() or (it.get("text") or "").strip()
        name = name.strip()

        seen.add(pdp)
        products.append({"product_name": name, "product_url": pdp})
        if len(products) >= max_products:
            break

    return products


# ------------------------------------------------------------
# PDP image extraction (product-only)
# ------------------------------------------------------------
def images_from_jsonld_product(html: str, base_url: str) -> list[str]:
    soup = BeautifulSoup(html, "html.parser")
    urls: list[str] = []

    for tag in soup.find_all("script", {"type": "application/ld+json"}):
        txt = tag.get_text(strip=True)
        if not txt:
            continue
        try:
            obj = json.loads(txt)
        except Exception:
            continue

        nodes = obj if isinstance(obj, list) else [obj]
        i = 0
        while i < len(nodes):
            n = nodes[i]
            i += 1
            if not isinstance(n, dict):
                continue

            if "@graph" in n and isinstance(n["@graph"], list):
                nodes.extend([x for x in n["@graph"] if isinstance(x, dict)])
                continue

            t = n.get("@type")
            is_product = ("Product" in t) if isinstance(t, list) else (t == "Product")
            if not is_product:
                continue

            img = n.get("image")
            if isinstance(img, str):
                urls.append(abs_url(base_url, img))
            elif isinstance(img, list):
                for x in img:
                    if isinstance(x, str):
                        urls.append(abs_url(base_url, x))
                    elif isinstance(x, dict) and x.get("url"):
                        urls.append(abs_url(base_url, x["url"]))
            elif isinstance(img, dict) and img.get("url"):
                urls.append(abs_url(base_url, img["url"]))

    out: list[str] = []
    seen: set[str] = set()
    for u in urls:
        u = normalize_image_url(u)
        key = image_dedupe_key(u)
        if key and key not in seen:
            seen.add(key)
            out.append(u)
    return out


def images_from_dom_gallery_playwright(pdp_url: str) -> list[str]:
    """
    Render PDP and extract product gallery images only.
    Priority:
      1) Swiper slides (Pas Normal uses swiper-slide for gallery)
      2) Common gallery/media containers
      3) Fallback to <main> but blacklist "also bought"/related areas
    """
    from playwright.sync_api import sync_playwright  # type: ignore

    with sync_playwright() as p:
        browser = _launch_browser(p)
        context = browser.new_context(user_agent=HEADERS["User-Agent"])

        # Speed: block heavy resources
        def _route(route, request):
            if request.resource_type in ("image", "media", "font"):
                return route.abort()
            return route.continue_()

        context.route("**/*", _route)

        page = context.new_page()
        page.goto(pdp_url, wait_until="domcontentloaded", timeout=90000)
        page.wait_for_timeout(400)

        js = r"""
        () => {
          const black = ['recommend','also','related','upsell','cross','carousel','brand','logo','footer','header','nav','newsletter','review','rating'];
          function isBad(el){
            let cur = el;
            for(let i=0;i<12 && cur;i++){
              const s = ((cur.id||'')+' '+(cur.className||'')).toLowerCase();
              if(black.some(b => s.includes(b))) return true;
              cur = cur.parentElement;
            }
            return false;
          }
          function pickUrl(img){
            const src = img.getAttribute('src') || '';
            if(src) return src;
            const srcset = img.getAttribute('srcset') || img.getAttribute('data-srcset') || '';
            if(srcset){
              const parts = srcset.split(',').map(x=>x.trim()).filter(Boolean);
              if(parts.length) return parts[parts.length-1].split(' ')[0].trim();
            }
            const dsrc = img.getAttribute('data-src') || '';
            if(dsrc) return dsrc;
            return '';
          }
          function okSize(img){
            const w = parseInt(img.getAttribute('width')||'0',10);
            const h = parseInt(img.getAttribute('height')||'0',10);
            if((w && w < 250) || (h && h < 250)) return false;
            return true;
          }

          const urls = [];
          const seen = new Set();

          // 1) Swiper gallery (preferred when present)
          // Many PDPs use Swiper both for product media and for recommendations.
          // We pick the *best* swiper container by scoring large images.
          const mainEl = document.querySelector('main') || document;
          const allSwiperImgs = Array.from(mainEl.querySelectorAll('.swiper-slide img'));
          if(allSwiperImgs.length){
            const groups = new Map(); // root -> {score, urls}
            function rootFor(img){
              return img.closest('.swiper') || img.closest('.swiper-wrapper') || img.closest('[class*="swiper"]') || img.parentElement;
            }
            for(const img of allSwiperImgs){
              if(isBad(img)) continue;
              if(!okSize(img)) continue;
              const u = pickUrl(img);
              if(!u) continue;
              const key = u.split('?')[0];

              const root = rootFor(img);
              if(!root) continue;
              if(!groups.has(root)) groups.set(root, {score: 0, urls: [], seen: new Set()});
              const g = groups.get(root);

              if(g.seen.has(key)) continue;
              g.seen.add(key);

              // score by declared width/height when available (bigger usually = product gallery)
              const w = parseInt(img.getAttribute('width')||'0',10);
              const h = parseInt(img.getAttribute('height')||'0',10);
              const area = (w>0 && h>0) ? (w*h) : 1;
              g.score += area;
              g.urls.push(u);
            }

            // pick best group by score, then by count
            let best = null;
            for(const g of groups.values()){
              if(!best) best = g;
              else if(g.score > best.score) best = g;
              else if(g.score === best.score && g.urls.length > best.urls.length) best = g;
            }
            if(best && best.urls.length){
              for(const u of best.urls){
                const key = u.split('?')[0];
                if(seen.has(key)) continue;
                seen.add(key);
                urls.push(u);
              }
              if(urls.length >= 2) return urls;
            }
          }

          // 2) Common gallery containers
          const selectors = [
            '[data-product-media]',
            '[data-product-gallery]',
            '[data-testid*="gallery"]',
            '[data-testid*="product-media"]',
            '[class*="ProductGallery"]',
            '[class*="product-gallery"]',
            '[class*="gallery"]',
            '[class*="Gallery"]',
            '[class*="media"]',
            '[class*="Media"]'
          ];
          let root = null;
          for(const sel of selectors){
            const el = document.querySelector(sel);
            if(el){ root = el; break; }
          }
          if(!root) root = document.querySelector('main') || document;

          const imgs = Array.from(root.querySelectorAll('img'));
          for(const img of imgs){
            if(isBad(img)) continue;
            if(!okSize(img)) continue;
            const u = pickUrl(img);
            if(!u) continue;
            const key = u.split('?')[0];
            if(seen.has(key)) continue;
            seen.add(key);
            urls.push(u);
          }
          return urls;
        }
        """
        urls = page.evaluate(js) or []
        browser.close()

    out: list[str] = []
    seen: set[str] = set()
    for u in urls:
        au = urljoin(pdp_url, u)
        au = normalize_image_url(au)
        key = image_dedupe_key(au)
        if key and key not in seen:
            seen.add(key)
            out.append(au)
    return out


def extract_product_images(pdp_url: str) -> list[str]:
    """
    Product-only images:
      1) JSON-LD Product
      2) Rendered DOM gallery (swiper/gallery) via Playwright
    """
    html = ""
    try:
        html = http_get(pdp_url).text
    except Exception:
        html = ""

    imgs = images_from_jsonld_product(html, pdp_url) if html else []

    if len(imgs) < 2:
        imgs = images_from_dom_gallery_playwright(pdp_url)

    # Final dedupe
    out: list[str] = []
    seen: set[str] = set()
    for u in imgs:
        u = normalize_image_url(u)
        key = image_dedupe_key(u)
        if key and key not in seen:
            seen.add(key)
            out.append(u)
    return out




st.set_page_config(page_title="Standardix", layout="wide")

st.title("Standardix – Outils eCommerce")

tool = st.sidebar.radio(
    "Choisissez un outil :",
    [
        "Standardiser les attributs",
        "Générer des descriptions courtes",
        "Extraire les images produits",
    ],
)

# --------------------------------------------------
# OUTIL 1 – STANDARDISATION DES ATTRIBUTS
# --------------------------------------------------
if tool == "Standardiser les attributs":
    st.header("🧩 Standardisation des attributs")

    uploaded_products = st.file_uploader(
        "Déposez votre fichier fournisseur (CSV ou Excel)",
        type=["csv", "xlsx", "xls"],
        key="products_std",
    )
    uploaded_mapping = st.file_uploader(
        "Déposez votre fichier de mapping (CSV ou Excel)",
        type=["csv", "xlsx", "xls"],
        key="mapping_std",
    )

    # --- Bouton en haut + placeholders juste en dessous ---
    start_standardization = st.button("Lancer la standardisation")
    status_placeholder = st.empty()
    download_placeholder = st.empty()

    # --- Texte & options de standardisation des mesures ---
    st.markdown(
        """
**Standardisation des mesures (pouces / cm)**  

Les mesures, s'il y en a à votre fichier, seront standardisées par défaut en **pouces en fraction (cm)**,  
par exemple : `1-1/4 po (3,18 cm)`.
        """
    )

    show_advanced_measures = st.checkbox(
        "Afficher les options avancées pour les mesures (pouces / centimètres)"
    )

    # Valeurs par défaut : fraction + 2 décimales + pouces + cm
    measure_options = {
        "mode_format": "fraction",   # 'fraction' ou 'decimale'
        "dec_places": 2,             # nb de décimales pour cm / pouces décimaux
        "add_unit": True,
        "unit_final": "les deux",    # pouces + cm par défaut
    }

    if show_advanced_measures:
        format_pouces = st.selectbox(
            "Format des pouces",
            options=[
                "Fraction (ex. 1-1/4)",
                "Décimal (ex. 1,25)",
            ],
            index=0,
        )
        if "Décimal" in format_pouces:
            measure_options["mode_format"] = "decimale"
        else:
            measure_options["mode_format"] = "fraction"

        dec_places = st.number_input(
            "Nombre de chiffres après la virgule pour les cm et les pouces en décimales",
            min_value=0,
            max_value=4,
            value=2,
            step=1,
        )
        measure_options["dec_places"] = int(dec_places)

        # Unité finale : pouces / cm / les deux
        unite_finale = st.selectbox(
            "Unité finale des mesures",
            options=[
                "Pouces + centimètres",
                "Seulement pouces",
                "Seulement centimètres",
            ],
            index=0,
        )
        if "Seulement pouces" in unite_finale:
            measure_options["unit_final"] = "in"
        elif "Seulement centimètres" in unite_finale:
            measure_options["unit_final"] = "cm"
        else:
            measure_options["unit_final"] = "les deux"

    if start_standardization:
        if uploaded_products and uploaded_mapping:

            with st.spinner("Standardisation en cours..."):
                # 1) Standardisation → DataFrames EN / FR
                df_en, df_fr = standardix(
                    uploaded_products,
                    uploaded_mapping,
                    measure_options=measure_options,
                )

                # 2) Lire le fichier fournisseur pour récupérer l'ordre initial
                df_products = read_table(uploaded_products)
                original_cols = list(df_products.columns)

                # 3) Réordonner les colonnes standardisées :
                #    -> elles suivent l'ordre des colonnes d'origine
                std_cols_en = []
                for col in original_cols:
                    cand = f"{col}_standard_en"
                    if cand in df_en.columns:
                        std_cols_en.append(cand)
                df_en = df_en[original_cols + std_cols_en]

                std_cols_fr = []
                for col in original_cols:
                    cand = f"{col}_standard_fr"
                    if cand in df_fr.columns:
                        std_cols_fr.append(cand)
                df_fr = df_fr[original_cols + std_cols_fr]

                # 4) Écriture dans un Excel en mémoire
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    df_en.to_excel(writer, sheet_name="EN", index=False)
                    df_fr.to_excel(writer, sheet_name="FR", index=False)

                buffer.seek(0)

                # 5) Recharger le fichier pour colorer les en-têtes
                wb = load_workbook(buffer)

                green_fill = PatternFill(start_color="00C6EFCE", end_color="00C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="00FFC7CE", end_color="00FFC7CE", fill_type="solid")

                # Colonnes qui ne doivent jamais être en rouge
                never_red = {"sku", "Short Description"}

                # ---------- FEUILLE EN ----------
                ws_en = wb["EN"]

                for col_idx, col_name in enumerate(df_en.columns, start=1):

                    # 1) Colonne standardisée → VERT
                    if col_name.endswith("_standard_en"):
                        ws_en.cell(row=1, column=col_idx).fill = green_fill
                        continue

                    # 2) Colonne d’origine → ROUGE SEULEMENT SI elle n’a pas été standardisée
                    if col_name in original_cols and col_name not in never_red:
                        std_version = f"{col_name}_standard_en"
                        if std_version not in df_en.columns:
                            ws_en.cell(row=1, column=col_idx).fill = red_fill

                # ---------- FEUILLE FR ----------
                ws_fr = wb["FR"]

                for col_idx, col_name in enumerate(df_fr.columns, start=1):

                    # 1) Colonne standardisée → VERT
                    if col_name.endswith("_standard_fr"):
                        ws_fr.cell(row=1, column=col_idx).fill = green_fill
                        continue

                    # 2) Colonne d’origine → ROUGE SEULEMENT SI elle n’a pas été standardisée
                    if col_name in original_cols and col_name not in never_red:
                        std_version = f"{col_name}_standard_fr"
                        if std_version not in df_fr.columns:
                            ws_fr.cell(row=1, column=col_idx).fill = red_fill

                # 6) Sauvegarde finale
                output = BytesIO()
                wb.save(output)
                output.seek(0)

            # ✅ Message + bouton apparaissent juste sous le bouton
            status_placeholder.success("✅ Standardisation terminée. Vous pouvez télécharger le fichier.")

            download_placeholder.download_button(
                "📥 Télécharger le fichier standardisé (Excel)",
                data=output,
                file_name="products_standardized.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        else:
            status_placeholder.error("Veuillez téléverser les deux fichiers (fournisseur et mapping).")

# --------------------------------------------------
# OUTIL 2 – GÉNÉRATION DES DESCRIPTIONS COURTES
# --------------------------------------------------
elif tool == "Générer des descriptions courtes":
    st.header("✏️ Générer des descriptions courtes")

    st.markdown(
        """
        1. Téléverse le **fichier standardisé** (sorti de l'outil précédent, avec les onglets EN / FR).  
        2. Téléverse le **fichier de recettes** (`short_description_recipes.xlsx`, avec Recipe_EN et Recipe_FR).  
        3. Clique sur le bouton pour générer un Excel avec les courtes descriptions EN / FR.
        """
    )

    uploaded_standardized = st.file_uploader(
        "Fichier standardisé (Excel, avec onglets EN et FR)",
        type=["xlsx", "xls"],
        key="standardized_shortdesc",
    )

    uploaded_recipes = st.file_uploader(
        "Fichier de recettes (short_description_recipes.xlsx)",
        type=["xlsx", "xls"],
        key="recipes_shortdesc",
    )

    if st.button("Générer les descriptions courtes"):
        if not uploaded_standardized or not uploaded_recipes:
            st.error("Merci de téléverser **les 2 fichiers** (standardisé + recettes).")
        else:
            try:
                # ----- Lecture des fichiers uploadés -----
                std_sheets = pd.read_excel(
                    uploaded_standardized,
                    sheet_name=[SHEET_EN, SHEET_FR],
                )
                en_std = std_sheets[SHEET_EN]
                fr_std = std_sheets[SHEET_FR]

                recipe_sheets = pd.read_excel(
                    uploaded_recipes,
                    sheet_name=[RECIPE_SHEET_EN, RECIPE_SHEET_FR],
                )
                recipes_en_raw = recipe_sheets[RECIPE_SHEET_EN]
                recipes_fr_raw = recipe_sheets[RECIPE_SHEET_FR]

                recipes_en = load_recipes(recipes_en_raw, RECIPE_SHEET_EN)
                recipes_fr = load_recipes(recipes_fr_raw, RECIPE_SHEET_FR)

                en_out = process_language(en_std, recipes_en, "EN")
                fr_out = process_language(fr_std, recipes_fr, "FR")

                # ----- Écriture dans un Excel en mémoire -----
                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    en_out.to_excel(writer, sheet_name="EN", index=False)
                    fr_out.to_excel(writer, sheet_name="FR", index=False)

                output.seek(0)

                # ----- Coloration des en-têtes -----
                wb = load_workbook(output)

                green_fill = PatternFill(
                    start_color="00C6EFCE",
                    end_color="00C6EFCE",
                    fill_type="solid",
                )
                red_fill = PatternFill(
                    start_color="00FFC7CE",
                    end_color="00FFC7CE",
                    fill_type="solid",
                )

                for sheet_name, df in [("EN", en_out), ("FR", fr_out)]:
                    ws = wb[sheet_name]
                    cols = list(df.columns)
                    for col_idx, col_name in enumerate(cols, start=1):
                        if col_name == COL_SHORT_DESC:  # short_description (nouvelle)
                            ws.cell(row=1, column=col_idx).fill = green_fill
                        elif col_name == ORIG_DESC_COL:  # Short Description (origine)
                            ws.cell(row=1, column=col_idx).fill = red_fill

                # ----- Sauvegarde finale -----
                final_output = BytesIO()
                wb.save(final_output)
                final_output.seek(0)

                st.success("✅ Descriptions courtes générées.")
                st.download_button(
                    "📥 Télécharger le fichier avec descriptions courtes",
                    data=final_output,
                    file_name="products_with_short_description.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            except Exception as e:
                st.error(f"Une erreur est survenue : {e}")

# ---------------- Tool 3
else:
    st.header("🖼️ Extraire les images produits")

    authorized = st.checkbox(
        "Je confirme disposer de l’autorisation du fournisseur pour extraire et utiliser les images de ses produits."
    )

    plp_url = st.text_input("URL de la page source des produits", placeholder="https://exemple.com/collections/produits")
    st.caption(
        "Inscrivez l’URL de la page fournisseur contenant la liste des produits (ex. page de collection ou catégorie, PLP)."
    )

    max_products = st.number_input("Nombre maximum de produits à traiter", min_value=1, max_value=300, value=150, step=1)

    if st.button("Extraire", disabled=(not authorized)):
        if not plp_url or not plp_url.startswith(("http://", "https://")):
            st.error("Veuillez fournir une URL valide (http/https).")
            st.stop()

        # Progress: X/Y only + message by ratio
        progress = st.progress(0)
        status = st.empty()

        def phase_message(done: int, total_items: int) -> str:
            if total_items <= 0:
                return "Traitement…"
            ratio = done / total_items
            if ratio <= 0.25:
                return "Récupérer les nom des produits"
            elif ratio <= 0.50:
                return "Récupérer les URL des produits"
            elif ratio <= 0.75:
                return "Récupérer les url des images principales"
            else:
                return "Récupérer les url des images additionnelles"

        def update(done: int, total_items: int):
            pct = int((done / total_items) * 100) if total_items else 0
            progress.progress(min(max(pct, 0), 100))
            status.text(f"{done}/{total_items} — {phase_message(done, total_items)}")

        requested_total = int(max_products)

        # Step A: product list (requests first, then Playwright)
        products: list[dict] = []
        try:
            products = plp_products_requests(plp_url, max_products=requested_total)
        except Exception:
            products = []

        if not products:
            try:
                products = plp_products_playwright(plp_url, max_products=requested_total)
            except Exception as e:
                st.error(
                    "Impossible de récupérer la liste des produits sur cette page. "
                    "Il se peut qu'elle charge le contenu via JavaScript ou bloque l'automatisation."
                )
                st.error(f"Une erreur est survenue : {e}")
                st.stop()

        products = products[:requested_total]
        total_items = len(products)
        if total_items == 0:
            st.error("Aucun produit détecté.")
            st.stop()

        update(0, total_items)

        # Step B: iterate products and extract product-only images
        rows: list[dict] = []
        for idx, pr in enumerate(products, start=1):
            product_name = (pr.get("product_name") or "").strip()
            product_url = (pr.get("product_url") or "").strip()

            # fallback name
            if not product_name:
                product_name = product_url.rsplit("/", 1)[-1].replace("-", " ").strip()

            # Extract images
            try:
                img_urls = extract_product_images(product_url)
            except Exception:
                img_urls = []

            # Shopify: one row per image
            alt_text = product_name
            for pos, img in enumerate(img_urls, start=1):
                rows.append(
                    {
                        "Product name": product_name,
                        "Product URL": product_url,
                        "Product image URL": img,
                        "Image position": pos,
                        "Image alt text": alt_text,
                    }
                )

            update(idx, total_items)

        if not rows:
            st.error(
                "Aucune image produit trouvée. (Le site peut bloquer l'accès ou les images sont rendues uniquement via JavaScript)"
            )
            st.stop()

        df = pd.DataFrame(
            rows,
            columns=["Product name", "Product URL", "Product image URL", "Image position", "Image alt text"],
        )

        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Images")

        st.download_button(
            "Télécharger le fichier Excel",
            data=buffer.getvalue(),
            file_name="product_images.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
